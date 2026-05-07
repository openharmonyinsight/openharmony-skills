#!/usr/bin/env python3
"""
Excel文件读取工具

用于读取覆盖率报告Excel文件并转换为可处理的格式
支持读取整个Excel工作表的内容
"""

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

import sys
import json
from typing import List, Tuple, Any


def read_excel_file(file_path: str) -> List[Tuple[Any, ...]]:
    """
    读取Excel文件的所有行数据

    Args:
        file_path: Excel文件路径

    Returns:
        包含所有行数据的列表，每行是一个元组
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        rows_data = []

        for row in ws.iter_rows(values_only=True):
            rows_data.append(row)

        return rows_data
    except FileNotFoundError:
        print(f"错误: 文件未找到 - {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"错误: 读取Excel文件失败 - {e}")
        sys.exit(1)


def read_excel_as_csv(file_path: str) -> str:
    """
    读取Excel文件并转换为CSV格式字符串

    Args:
        file_path: Excel文件路径

    Returns:
        CSV格式的字符串
    """
    import csv
    import io

    rows_data = read_excel_file(file_path)

    output = io.StringIO()
    writer = csv.writer(output)

    for row in rows_data:
        writer.writerow(row)

    return output.getvalue()


def read_excel_as_json(file_path: str) -> List[dict]:
    """
    读取Excel文件并转换为JSON格式

    Args:
        file_path: Excel文件路径

    Returns:
        包含所有行数据的字典列表，第一行作为键
    """
    rows_data = read_excel_file(file_path)

    if not rows_data:
        return []

    headers = [str(h) if h is not None else "" for h in rows_data[0]]
    result = []

    for row in rows_data[1:]:
        row_dict = {}
        for i, value in enumerate(row):
            key = headers[i] if i < len(headers) else f"column_{i}"
            row_dict[key] = value
        result.append(row_dict)

    return result


def print_excel_content(file_path: str, format: str = "table") -> None:
    """
    打印Excel文件内容

    Args:
        file_path: Excel文件路径
        format: 输出格式 (table, csv, json)
    """
    rows_data = read_excel_file(file_path)

    if format == "csv":
        print(read_excel_as_csv(file_path))
    elif format == "json":
        print(json.dumps(read_excel_as_json(file_path), indent=2, ensure_ascii=False))
    else:  # table格式 (默认)
        for row in rows_data:
            print(row)


def rows_to_csv(rows_data):
    import csv
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows_data:
        writer.writerow(row)
    return output.getvalue()


def rows_to_json(rows_data):
    if not rows_data:
        return []
    headers = [str(h) if h is not None else "" for h in rows_data[0]]
    result = []
    for row in rows_data[1:]:
        row_dict = {}
        for i, value in enumerate(row):
            key = headers[i] if i < len(headers) else f"column_{i}"
            row_dict[key] = value
        result.append(row_dict)
    return result


def main():
    import argparse

    parser = argparse.ArgumentParser(description="读取Excel文件内容")
    parser.add_argument("file", help="Excel文件路径")
    parser.add_argument(
        "-f", "--format",
        choices=["table", "csv", "json"],
        default="table",
        help="输出格式 (默认: table)"
    )
    parser.add_argument(
        "-o", "--output",
        help="输出文件路径 (可选，默认输出到控制台)"
    )
    parser.add_argument(
        "-s", "--sheet",
        help="工作表名称或索引 (默认: 最后一个工作表，通常含覆盖率数据)",
        default=None
    )

    args = parser.parse_args()

    try:
        wb = openpyxl.load_workbook(args.file)
    except FileNotFoundError:
        print(f"错误: 文件未找到 - {args.file}")
        sys.exit(1)
    except Exception as e:
        print(f"错误: 读取Excel文件失败 - {e}")
        sys.exit(1)

    if args.sheet is not None:
        if args.sheet.isdigit():
            ws = wb.worksheets[int(args.sheet)]
        else:
            ws = wb[args.sheet]
    else:
        ws = wb.worksheets[-1]

    rows_data = list(ws.iter_rows(values_only=True))

    if args.format == "csv":
        content = rows_to_csv(rows_data)
    elif args.format == "json":
        content = json.dumps(rows_to_json(rows_data), indent=2, ensure_ascii=False)
    else:
        lines = [str(row) for row in rows_data]
        content = "\n".join(lines)

    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"内容已保存到: {args.output}")
    else:
        print(content)

    wb.close()


if __name__ == "__main__":
    main()
