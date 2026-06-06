#!/usr/bin/env python3
"""Extract uncovered APIs from APICoverageDetector coverage reports.

Supports both Excel (.xlsx) and CSV formats. Evaluates 8 orthogonal
dimensions (AQ-AX) to determine coverage status, and outputs two files:
  - uncovered_apis_{timestamp}.json  — APIs with uncovered dimensions
  - manual_confirm_{timestamp}.json  — APIs requiring manual confirmation

Usage:
    python extract_uncovered.py --subsystem "ArkUI开发框架"
    python extract_uncovered.py --dts-file "component\\text.d.ts"
    python extract_uncovered.py --class-name TextAttribute --interface-name fontFeature
    python extract_uncovered.py --kit ArkUI --iter-phase 2
"""

import glob
import os
import json
import csv
import argparse
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_ROOT = os.path.join(SCRIPT_DIR, '..')

DIMENSIONS = ['call', 'param', 'param_spec', 'return_value', 'error_code', 'permission', 'stage', 'meta']

COVERED_VALUES = {'工具检测已覆盖', '已覆盖', '本地覆盖', '手工用例'}
NOT_APPLICABLE = '不涉及'
MANUAL_CONFIRM_DESC = '工具暂不能识别返回值类型，请人工确认'

SKIPPED_TYPES = {'Import', 'TypeAlias', 'EnumValue'}

# ArkUI component call coverage heuristic:
# For ArkUI component APIs, the "call" dimension requires invocation in a test function,
# but the Inspector pattern (Demo page + getInspectorByKey) only has calls in page build().
# This causes false "uncovered" for call even when param/return_value are covered.
# These APIs should be moved to manual_confirm instead of appearing as "needs test generation".
ARKUI_COMPONENT_CALL_HEURISTIC = {
    'subsystem': 'ArkUI开发框架',
    'file_path_prefix': 'component',  # file_path starts with "component"
    'call_reason': 'ArkUI组件API使用Inspector模式测试，call维度仅在page build()中调用，'
                    '工具无法识别为测试覆盖，需人工确认',
}

COL_FORMAT_24 = {
    'module': 0, 'class': 1, 'method': 2, 'func': 3, 'type': 4,
    'start_version': 5, 'latest_version': 6, 'deprecate_version': 7,
    'syscap': 8, 'error_codes': 9, 'is_system_api': 10,
    'model_constraint': 11, 'permission': 12,
    'kit': 19, 'file_path': 20, 'subsystem': 21,
    'parent_type': 22, 'is_optional': 23,
}

COL_FORMAT_27 = {
    'module': 0, 'class': 1, 'method': 2, 'func': 3, 'type': 4,
    'start_version': 5, 'deprecate_version': 7,
    'error_codes': 9, 'model_constraint': 11,
    'kit': 17, 'subsystem': 19,
    'param_optional': 20, 'params': 21, 'param_range': 22,
    'is_promise': 23, 'is_covered': 26,
    'uncovered_err': 28,
    'permission_covered': 29, 'permission_uncovered': 30,
    'param_coverage': 31, 'param_uncovered_desc': 32,
    'param_coverage_note': 33,
    'return_value_coverage': 34, 'return_value_status': 35,
    'return_value_uncovered_desc': 36,
    'return_value_uncovered_note': 37,
}

XLSX_DIMENSION_HEADER_PATTERNS = {
    'call': ['调用覆盖白名单', '调用覆盖'],
    'param': ['入参覆盖白名单', '入参覆盖'],
    'param_spec': ['参数规格覆盖白名单', '参数规格覆盖'],
    'return_value': ['返回值覆盖白名单', '返回值覆盖'],
    'error_code': ['错误码覆盖白名单', '错误码覆盖'],
    'permission': ['权限覆盖白名单', '权限覆盖'],
    'stage': ['stagemode标签覆盖白名单', 'stagemode标签覆盖'],
    'meta': ['元服务覆盖白名单', '元服务覆盖'],
}

XLSX_CORE_HEADER_MAP = {
    'module': '模块名',
    'class': '类名',
    'method': '方法名',
    'func': '函数',
    'type': '类型',
    'start_version': '起始版本',
    'deprecate_version': '废弃版本',
    'error_codes': '错误码',
    'model_constraint': '模型限制',
    'kit': 'kit',
    'file_path': '文件路径',
    'subsystem': '子系统',
    'interface_covered_status': '接口覆盖是否已满足',
}

XLSX_ERR_DESC_HEADERS = {
    'return_value': ['返回值覆盖错误说明', '返回值覆盖描述'],
}

LEGACY_XLSX_DIMENSION_COLS = {
    'call': 'AQ', 'param': 'AR', 'param_spec': 'AS',
    'return_value': 'AT', 'error_code': 'AU', 'permission': 'AV',
    'stage': 'AW', 'meta': 'AX',
}

LEGACY_XLSX_CORE_COLS = {
    'module': 'A', 'class': 'B', 'method': 'C', 'func': 'D', 'type': 'E',
    'start_version': 'F', 'deprecate_version': 'H', 'error_codes': 'J',
    'model_constraint': 'L', 'kit': 'R', 'file_path': 'U', 'subsystem': 'V',
    'interface_covered_status': 'AP',
}

COL_LETTER_TO_INDEX = {}
for _i, _c in enumerate('ABCDEFGHIJKLMNOPQRSTUVWXYZ'):
    COL_LETTER_TO_INDEX[_c] = _i
    if _i < 26:
        for _j, _c2 in enumerate('ABCDEFGHIJKLMNOPQRSTUVWXYZ'):
            COL_LETTER_TO_INDEX[_c + _c2] = 26 * (_i + 1) + _j


def col_to_index(col_str):
    return COL_LETTER_TO_INDEX.get(col_str.upper())


def build_header_index(headers):
    idx_map = {}
    for i, h in enumerate(headers):
        h_str = str(h).strip() if h else ''
        if h_str:
            idx_map[h_str] = i
    return idx_map


def find_col_by_headers(header_index, candidates, default=None):
    for name in candidates:
        if name in header_index:
            return header_index[name]
    return default


def resolve_core_cols(headers):
    header_index = build_header_index(headers)
    col_map = {}
    for key, header_name in XLSX_CORE_HEADER_MAP.items():
        if header_name in header_index:
            col_map[key] = header_index[header_name]
        elif key in LEGACY_XLSX_CORE_COLS:
            idx = col_to_index(LEGACY_XLSX_CORE_COLS[key])
            if idx is not None:
                col_map[key] = idx
    return col_map, header_index


def resolve_dimension_cols(headers):
    header_index = build_header_index(headers)
    dim_cols = {}
    for dim_name, patterns in XLSX_DIMENSION_HEADER_PATTERNS.items():
        idx = find_col_by_headers(header_index, patterns)
        if idx is not None:
            dim_cols[dim_name] = idx
        elif dim_name in LEGACY_XLSX_DIMENSION_COLS:
            idx = col_to_index(LEGACY_XLSX_DIMENSION_COLS[dim_name])
            if idx is not None:
                dim_cols[dim_name] = idx
    return dim_cols


def load_config():
    config_path = os.path.join(SKILL_ROOT, '.oh-xts-config.json')
    if os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
            ets_version = config.get('ets_version', ['ets1.1'])
            if isinstance(ets_version, str):
                ets_version = [v.strip() for v in ets_version.split(',')]
            scan_tool_root = config.get('scan_tool_root', '')
            if not scan_tool_root or not os.path.isdir(scan_tool_root):
                scan_tool_root = ''
            return [v.strip() for v in ets_version if v.strip()], scan_tool_root
    return ['ets1.1'], ''


def safe_get(vals, index, default=''):
    if index < len(vals) and vals[index] is not None and str(vals[index]).strip() not in ('', 'None'):
        return str(vals[index]).strip()
    return default


def detect_csv_col_format(vals):
    if len(vals) >= 27:
        return '27+'
    elif len(vals) >= 20:
        return '24'
    return None


def find_latest_excel(results_dir, ets_version, iter_phase=None):
    search_dirs = [results_dir]
    if iter_phase:
        iter_dir = os.path.join(results_dir, f'iter-{iter_phase}')
        if os.path.isdir(iter_dir):
            search_dirs.insert(0, iter_dir)

    for search_base in search_dirs:
        for pattern_name in ['sdk_result_开源*.xlsx', 'sdk_result*.xlsx']:
            found = sorted(glob.glob(os.path.join(search_base, pattern_name)), key=os.path.getmtime, reverse=True)
            if found:
                return found[0]

    return None


def find_latest_csv(results_dir, ets_version, iter_phase=None, file_prefix='before_generation'):
    prefixes = [file_prefix, 'all_collect', 'collect']
    for prefix in prefixes:
        if iter_phase:
            pattern = os.path.join(results_dir, f'iter-{iter_phase}', f'{prefix}_{ets_version}_*.csv')
        else:
            pattern = os.path.join(results_dir, f'{prefix}_{ets_version}_*.csv')
        files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
        if files:
            return files[0]

        if iter_phase:
            pattern = os.path.join(results_dir, f'iter-{iter_phase}', f'{prefix}_*.csv')
        else:
            pattern = os.path.join(results_dir, f'{prefix}_*.csv')
        files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
        if files:
            return files[0]
    return None


def judge_dimension(whitelist_value, err_desc=''):
    wl = str(whitelist_value).strip() if whitelist_value is not None else ''
    if wl in COVERED_VALUES:
        return '已覆盖'
    if wl == NOT_APPLICABLE:
        return '不涉及'
    if MANUAL_CONFIRM_DESC in str(err_desc):
        return '需人工确认'
    if wl == '':
        return '未覆盖'
    if wl in ('澄清无法覆盖', '单独代码仓'):
        return '不涉及'
    return '已覆盖'


def parse_excel_dimensions(vals, dim_cols, header_index):
    coverage = {}
    for dim_name, col_idx in dim_cols.items():
        if col_idx < len(vals):
            cell_val = vals[col_idx]
        else:
            cell_val = None
        err_desc = ''
        err_patterns = XLSX_ERR_DESC_HEADERS.get(dim_name, [])
        for ep in err_patterns:
            if ep in header_index:
                ei = header_index[ep]
                if ei < len(vals) and vals[ei]:
                    err_desc = str(vals[ei]).strip()
                    break
        wl = str(cell_val).strip() if cell_val is not None else ''
        status = judge_dimension(wl, err_desc)
        if status == '未覆盖':
            coverage[dim_name] = {'status': '未覆盖'}
        elif status == '需人工确认':
            coverage[dim_name] = {'status': '需人工确认', 'err_desc': err_desc}
    return coverage


def _is_arkui_component_call_heuristic_match(entry, dim_name, header_index=None, raw_vals=None, col_map=None):
    """Check if an uncovered dimension qualifies for ArkUI component call heuristic.

    Matches when ALL of:
    - subsystem == "ArkUI开发框架"
    - file_path starts with "component"
    - dimension is 'call'
    - The API has call_file_path (col 41) not empty (called somewhere)
    - The API has call_whitelist (col 44) empty (not yet whitelisted)

    When using raw Excel data (raw_vals provided), checks the actual Excel columns.
    When only entry dict is available, uses entry fields (less precise but still useful).
    """
    if dim_name != 'call':
        return False

    sub = entry.get('subsystem', '')
    fp = entry.get('file_path', '')

    if sub != ARKUI_COMPONENT_CALL_HEURISTIC['subsystem']:
        return False
    if not fp.lower().startswith(ARKUI_COMPONENT_CALL_HEURISTIC['file_path_prefix']):
        return False

    # If we have raw Excel data, check columns 41 and 44
    if raw_vals is not None and col_map is not None:
        # Build header_index from raw_vals would require headers; fall back to column indices
        # Col 41 = 调用文件路径, Col 44 = 调用覆盖白名单
        call_file = str(raw_vals[41]).strip() if len(raw_vals) > 41 and raw_vals[41] is not None else ''
        call_wl = str(raw_vals[44]).strip() if len(raw_vals) > 44 and raw_vals[44] is not None else ''
        has_call_file = len(call_file) > 0
        no_call_whitelist = call_wl in ('', 'None')
    else:
        # Fallback: if the entry has 'call' as uncovered and we know it's ArkUI component,
        # that's sufficient for the heuristic
        has_call_file = True  # Assume yes since it's showing up
        no_call_whitelist = True

    return has_call_file and no_call_whitelist


def _apply_arkui_component_heuristic(methods, interfaces, properties, manual_confirm_list,
                                      raw_rows=None, headers=None, col_map=None):
    """Post-process: move ArkUI component 'call' dimension from uncovered to manual_confirm.

    For APIs matching the ArkUI component call heuristic, the 'call' dimension is
    removed from the uncovered list and added to manual_confirm with a clear reason.
    Other uncovered dimensions (param, return_value, etc.) remain unchanged.
    """
    if not raw_rows or not headers:
        # Without raw Excel data, we can still apply a simplified version
        # by checking entry fields only
        pass

    def _process_list(entries, mc_list, entry_raw_map=None):
        processed = []
        for entry in entries:
            coverage = entry.get('coverage', {})
            if 'call' not in coverage:
                processed.append(entry)
                continue

            # Check heuristic
            sub = entry.get('subsystem', '')
            fp = entry.get('file_path', '')
            is_match = (
                sub == ARKUI_COMPONENT_CALL_HEURISTIC['subsystem'] and
                fp.lower().startswith(ARKUI_COMPONENT_CALL_HEURISTIC['file_path_prefix'])
            )

            if not is_match:
                processed.append(entry)
                continue

            # Check raw data for call_file and call_whitelist
            if entry_raw_map:
                entry_key = f"{entry.get('module', '')}|{entry.get('class', '')}|{entry.get('method', '')}|{entry.get('func', '')}"
                raw_vals = entry_raw_map.get(entry_key)
                if raw_vals is not None:
                    call_file = str(raw_vals[41]).strip() if len(raw_vals) > 41 and raw_vals[41] is not None else ''
                    call_wl = str(raw_vals[44]).strip() if len(raw_vals) > 44 and raw_vals[44] is not None else ''
                    if not call_file or call_wl not in ('', 'None'):
                        is_match = False

            if not is_match:
                processed.append(entry)
                continue

            # Move 'call' from coverage to manual_confirm
            remaining_coverage = {k: v for k, v in coverage.items() if k != 'call'}

            mc_entry = {
                'module': entry.get('module', ''),
                'class': entry.get('class', ''),
                'method': entry.get('method', ''),
                'func': entry.get('func', ''),
                'file_path': entry.get('file_path', ''),
                'subsystem': entry.get('subsystem', ''),
                'kit': entry.get('kit', ''),
                'call': {
                    'status': '需人工确认',
                    'err_desc': ARKUI_COMPONENT_CALL_HEURISTIC['call_reason'],
                },
            }
            mc_list.append(mc_entry)

            if remaining_coverage:
                entry['coverage'] = remaining_coverage
                processed.append(entry)
            # If only 'call' was uncovered, entry is removed from uncovered list

        return processed

    # Build raw data map for precise checking
    entry_raw_map = {}
    if raw_rows and headers and col_map:
        for vals in raw_rows:
            if len(vals) > max(col_map.values(), default=0):
                key = f"{str(vals[col_map.get('module', 0)] or '').strip()}|{str(vals[col_map.get('class', 1)] or '').strip()}|{str(vals[col_map.get('method', 2)] or '').strip()}|{str(vals[col_map.get('func', 3)] or '').strip()}"
                entry_raw_map[key] = vals

    methods = _process_list(methods, manual_confirm_list, entry_raw_map)
    interfaces = _process_list(interfaces, manual_confirm_list, entry_raw_map)
    properties = _process_list(properties, manual_confirm_list, entry_raw_map)

    return methods, interfaces, properties, manual_confirm_list


def parse_excel_file(file_path, subsystem=None, kit=None, dts_file=None, class_name=None, interface_name=None):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    target_sheet = None
    for sn in ['接口覆盖详情', '接口覆盖', '覆盖详情']:
        if sn in wb.sheetnames:
            target_sheet = wb[sn]
            break
    if target_sheet is None:
        target_sheet = wb.active
    ws = target_sheet

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if headers is None:
        wb.close()
        return [], [], [], []

    col_map, header_index = resolve_core_cols(headers)
    dim_cols = resolve_dimension_cols(headers)

    methods = []
    interfaces = []
    properties = []
    manual_confirm_list = []
    raw_rows = []  # Collect raw rows for ArkUI component heuristic

    for row in rows_iter:
        vals = list(row) if row else []
        raw_rows.append(vals)

        api_type = safe_get(vals, col_map.get('type', 4))
        if api_type in SKIPPED_TYPES:
            continue

        if subsystem:
            sv = safe_get(vals, col_map.get('subsystem', -1))
            if subsystem != sv:
                continue

        if kit:
            kv = safe_get(vals, col_map.get('kit', -1))
            if kit != kv:
                continue

        file_path_val = safe_get(vals, col_map.get('file_path', -1))
        if dts_file:
            if dts_file not in file_path_val:
                continue

        cls_val = safe_get(vals, col_map.get('class', 1))
        if class_name:
            if class_name not in cls_val:
                continue

        method_val = safe_get(vals, col_map.get('method', 2))
        if interface_name:
            if interface_name not in method_val:
                continue

        coverage = parse_excel_dimensions(vals, dim_cols, header_index)

        filtered_coverage = {k: v for k, v in coverage.items() if v['status'] == '未覆盖'}
        manual_confirm_dims = {k: v for k, v in coverage.items() if v['status'] == '需人工确认'}

        interface_covered_status = safe_get(vals, col_map.get('interface_covered_status', -1))

        if manual_confirm_dims:
            mc_entry = {
                'module': safe_get(vals, col_map.get('module', 0)),
                'class': cls_val,
                'method': method_val,
                'func': safe_get(vals, col_map.get('func', 3)),
                'file_path': file_path_val,
                'subsystem': safe_get(vals, col_map.get('subsystem', 21)),
                'kit': safe_get(vals, col_map.get('kit', 17)),
            }
            mc_entry.update(manual_confirm_dims)
            manual_confirm_list.append(mc_entry)

        if filtered_coverage:
            entry = {
                'module': safe_get(vals, col_map.get('module', 0)),
                'class': cls_val,
                'method': method_val,
                'type': api_type,
                'func': safe_get(vals, col_map.get('func', 3)),
                'kit': safe_get(vals, col_map.get('kit', 17)),
                'file_path': file_path_val,
                'subsystem': safe_get(vals, col_map.get('subsystem', 21)),
                'error_codes': safe_get(vals, col_map.get('error_codes', 9)),
                'start_version': safe_get(vals, col_map.get('start_version', 5)),
                'stage_label': safe_get(vals, col_map.get('model_constraint', 11)),
                'interface_covered_status': interface_covered_status,
                'coverage': filtered_coverage,
            }
            entry = {k: v for k, v in entry.items() if v or k in ('coverage',)}

            if api_type == 'Method':
                methods.append(entry)
            elif api_type == 'Interface':
                interfaces.append(entry)
            elif api_type == 'Property':
                properties.append(entry)

    wb.close()

    # Apply ArkUI component call coverage heuristic
    methods, interfaces, properties, manual_confirm_list = _apply_arkui_component_heuristic(
        methods, interfaces, properties, manual_confirm_list,
        raw_rows=raw_rows, headers=headers, col_map=col_map,
    )

    return methods, interfaces, properties, manual_confirm_list


def parse_csv_file(file_path, subsystem=None, kit=None, dts_file=None, class_name=None, interface_name=None):
    rows = []
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)

    if len(rows) < 2:
        return [], [], [], []

    sample_row = rows[1] if len(rows[1]) >= len(rows[0]) else rows[0]
    fmt = detect_csv_col_format(sample_row)
    if fmt is None:
        print(f"Warning: unrecognized CSV format ({len(sample_row)} columns) in {file_path}")
        return [], [], [], []

    col = COL_FORMAT_27 if fmt == '27+' else COL_FORMAT_24

    methods = []
    interfaces = []
    properties = []
    manual_confirm_list = []

    for row in rows[1:]:
        vals = list(row)

        api_type = safe_get(vals, col.get('type', 4))
        if api_type in SKIPPED_TYPES:
            continue

        subsystem_val = safe_get(vals, col.get('subsystem', 21))
        if subsystem and subsystem != subsystem_val:
            continue

        kit_val = safe_get(vals, col.get('kit', 19 if fmt == '24' else 17))
        if kit and kit != kit_val:
            continue

        file_path_val = safe_get(vals, col.get('file_path', 20))
        if dts_file and dts_file not in file_path_val:
            continue

        cls_val = safe_get(vals, col.get('class', 1))
        if class_name and class_name not in cls_val:
            continue

        method_val = safe_get(vals, col.get('method', 2))
        if interface_name and interface_name not in method_val:
            continue

        if fmt == '27+' and 'is_covered' in col:
            is_covered = safe_get(vals, col['is_covered'])
            if is_covered and is_covered.lower() not in ('', 'false', '0'):
                continue

        coverage = {}
        if fmt == '27+':
            if safe_get(vals, col.get('param_coverage', 31), '').lower() != 'true':
                coverage['param'] = {'status': '未覆盖'}
            rv_uncovered = safe_get(vals, col.get('return_value_uncovered_desc', 36))
            if rv_uncovered:
                coverage['return_value'] = {'status': '未覆盖', 'err_desc': rv_uncovered}
            uncovered_err = safe_get(vals, col.get('uncovered_err', 28))
            if uncovered_err:
                coverage['error_code'] = {'status': '未覆盖', 'err_desc': uncovered_err}
            perm_uncovered = safe_get(vals, col.get('permission_uncovered', 30))
            if perm_uncovered:
                coverage['permission'] = {'status': '未覆盖', 'err_desc': perm_uncovered}
        else:
            coverage['call'] = {'status': '未覆盖'}

        if not coverage:
            continue

        filtered_coverage = {k: v for k, v in coverage.items() if v['status'] == '未覆盖'}
        manual_confirm_dims = {k: v for k, v in coverage.items() if v['status'] == '需人工确认'}

        if manual_confirm_dims:
            mc_entry = {
                'module': safe_get(vals, col.get('module', 0)),
                'class': cls_val,
                'method': method_val,
                'func': safe_get(vals, col.get('func', 3)),
                'file_path': file_path_val,
                'subsystem': subsystem_val,
                'kit': kit_val,
            }
            mc_entry.update(manual_confirm_dims)
            manual_confirm_list.append(mc_entry)

        if not filtered_coverage:
            continue

        entry = {
            'module': safe_get(vals, col.get('module', 0)),
            'class': cls_val,
            'method': method_val,
            'type': api_type,
            'func': safe_get(vals, col.get('func', 3)),
            'kit': kit_val,
            'file_path': file_path_val,
            'subsystem': subsystem_val,
            'error_codes': safe_get(vals, col.get('error_codes', 9)),
            'start_version': safe_get(vals, col.get('start_version', 5)),
            'stage_label': safe_get(vals, col.get('model_constraint', 11)),
            'coverage': filtered_coverage,
        }
        entry = {k: v for k, v in entry.items() if v or k in ('coverage',)}

        if api_type == 'Method':
            methods.append(entry)
        elif api_type == 'Interface':
            interfaces.append(entry)
        elif api_type == 'Property':
            properties.append(entry)

    return methods, interfaces, properties, manual_confirm_list


def merge_results(all_results):
    merged_methods = []
    merged_interfaces = []
    merged_properties = []
    merged_manual = []
    for m, i, p, mc in all_results:
        merged_methods.extend(m)
        merged_interfaces.extend(i)
        merged_properties.extend(p)
        merged_manual.extend(mc)
    return merged_methods, merged_interfaces, merged_properties, merged_manual


def deduplicate_entries(entries):
    seen = {}
    for entry in entries:
        key = f"{entry.get('module', '')}|{entry.get('class', '')}|{entry.get('method', '')}|{entry.get('func', '')}"
        if key not in seen:
            seen[key] = entry
    return list(seen.values())


def main(subsystem=None, kit=None, dts_file=None, class_name=None, interface_name=None, module_name=None, iter_phase=1,
         task_subsystem=None, task_module=None, session_id=None):
    ets_versions, scan_tool_root = load_config()
    coverage_data_dir = os.path.join(SKILL_ROOT, '.coverage_data')
    task_dir = coverage_data_dir
    if task_subsystem and task_module:
        task_dir = os.path.join(task_dir, task_subsystem, task_module)
    if session_id:
        task_dir = os.path.join(task_dir, session_id)

    version_data = {}
    all_manual_confirm = []
    total_methods = 0
    total_interfaces = 0
    total_properties = 0
    total_manual = 0

    coverage_data_dir = os.path.join(SKILL_ROOT, '.coverage_data')

    for ets_version in ets_versions:
        all_results = []
        source_desc = None

        if scan_tool_root:
            results_dir = os.path.join(scan_tool_root, 'results', ets_version, 'open_source')
            xlsx_file = find_latest_excel(results_dir, ets_version, iter_phase)
            if xlsx_file:
                source_desc = f"Excel: {os.path.basename(xlsx_file)}"
                print(f"Using Excel: {xlsx_file} ({os.path.getsize(xlsx_file)} bytes)")
                result = parse_excel_file(
                    xlsx_file,
                    subsystem=subsystem,
                    kit=kit,
                    dts_file=dts_file,
                    class_name=class_name,
                    interface_name=interface_name if interface_name else module_name,
                )
                all_results.append(result)

        csv_candidates = []
        for prefix in [f'before_generation_{ets_version}', f'all_collect_{ets_version}', f'collect_{ets_version}',
                       'before_generation', 'all_collect', 'collect']:
            path = find_latest_csv(task_dir, ets_version, iter_phase=iter_phase, file_prefix=prefix)
            if path and path not in [c for c in csv_candidates]:
                csv_candidates.append(path)

        for csv_file in csv_candidates:
            if not source_desc:
                source_desc = f"CSV: {os.path.basename(csv_file)}"
            print(f"Using CSV: {csv_file} ({os.path.getsize(csv_file)} bytes)")
            result = parse_csv_file(
                csv_file,
                subsystem=subsystem,
                kit=kit,
                dts_file=dts_file,
                class_name=class_name,
                interface_name=interface_name if interface_name else module_name,
            )
            all_results.append(result)

        if not all_results:
            print(f"No usable data file found for {ets_version}")
            continue

        methods, interfaces, properties, manual_confirm_list = merge_results(all_results)
        methods = deduplicate_entries(methods)
        interfaces = deduplicate_entries(interfaces)
        properties = deduplicate_entries(properties)

        version_data[ets_version] = {
            'methods': methods,
            'interfaces': interfaces,
            'properties': properties,
        }

        all_manual_confirm.extend(manual_confirm_list)

        total_methods += len(methods)
        total_interfaces += len(interfaces)
        total_properties += len(properties)
        total_manual += len(manual_confirm_list)

        print(f"{ets_version}: {len(methods)} methods, {len(interfaces)} interfaces, {len(properties)} properties, {len(manual_confirm_list)} manual_confirm")

    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    output_dir = os.path.join(task_dir, f'iter-{iter_phase}')
    os.makedirs(output_dir, exist_ok=True)

    uncovered_output = {
        'ets1.1': version_data.get('ets1.1', {'methods': [], 'interfaces': [], 'properties': []}),
        'ets1.2': version_data.get('ets1.2', {'methods': [], 'interfaces': [], 'properties': []}),
        'metadata': {
            'ets_versions': ets_versions,
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'summary': {},
        },
    }

    for ver in ets_versions:
        vd = version_data.get(ver, {})
        uncovered_output['metadata']['summary'][ver] = {
            'methods': len(vd.get('methods', [])),
            'interfaces': len(vd.get('interfaces', [])),
            'properties': len(vd.get('properties', [])),
        }

    uncovered_output = {k: v for k, v in uncovered_output.items() if v}

    uncovered_path = os.path.join(output_dir, f'uncovered_apis_{timestamp}.json')
    with open(uncovered_path, 'w', encoding='utf-8') as fp:
        json.dump(uncovered_output, fp, indent=2, ensure_ascii=False)

    if all_manual_confirm:
        all_manual_confirm = deduplicate_entries(all_manual_confirm)
        manual_output = {
            'manual_confirm': all_manual_confirm,
            'count': len(all_manual_confirm),
        }
        manual_path = os.path.join(output_dir, f'manual_confirm_{timestamp}.json')
        with open(manual_path, 'w', encoding='utf-8') as fp:
            json.dump(manual_output, fp, indent=2, ensure_ascii=False)
    else:
        manual_path = None

    print(f"\n=== SUMMARY ===")
    print(f"Processing: {', '.join(ets_versions)}")
    print(f"Total uncovered APIs: {total_methods} methods, {total_interfaces} interfaces, {total_properties} properties")
    if total_manual:
        # Count heuristic-impacted entries
        heuristic_count = sum(1 for m in all_manual_confirm
                             if isinstance(m.get('call'), dict)
                             and ARKUI_COMPONENT_CALL_HEURISTIC['call_reason'][:20] in m.get('call', {}).get('err_desc', ''))
        print(f"Manual confirm: {total_manual} APIs (incl. {heuristic_count} from ArkUI component call heuristic)")
    print(f"Uncovered APIs saved to: {uncovered_path}")
    if manual_path:
        print(f"Manual confirm saved to: {manual_path}")

    return uncovered_path, manual_path


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Extract uncovered APIs from coverage reports (Excel/CSV)')
    parser.add_argument('--subsystem', type=str, default=None, help='Filter by subsystem name (exact match)')
    parser.add_argument('--kit', type=str, default=None, help='Filter by kit name (exact match)')
    parser.add_argument('--dts-file', type=str, default=None, help='Filter by d.ts file path (partial match)')
    parser.add_argument('--class-name', type=str, default=None, help='Filter by class name (partial match)')
    parser.add_argument('--interface-name', type=str, default=None, help='Filter by interface/method name (partial match)')
    parser.add_argument('--module-name', type=str, default=None, help='Filter by module name (partial match, alias for --interface-name)')
    parser.add_argument('--iter-phase', type=int, default=1, help='Iteration phase number (default: 1)')
    parser.add_argument('--task-subsystem', type=str, default=None, help='Task subsystem for output path (e.g. testfwk)')
    parser.add_argument('--task-module', type=str, default=None, help='Task module for output path (e.g. uitest)')
    parser.add_argument('--session-id', type=str, default=None, help='Session ID matching .task_summary/session_XXX (e.g. session_20260602_150300)')
    args = parser.parse_args()

    main(
        subsystem=args.subsystem,
        kit=args.kit,
        dts_file=args.dts_file,
        class_name=args.class_name,
        interface_name=args.interface_name,
        module_name=args.module_name,
        iter_phase=args.iter_phase,
        task_subsystem=args.task_subsystem,
        task_module=args.task_module,
        session_id=getattr(args, 'session_id', None),
    )
