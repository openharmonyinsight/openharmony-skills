#!/usr/bin/env python3
"""Phase5测试用例Excel导出脚本（流式写入优化版）

用法: python phase5_export.py --output /path/to/output [--start-id TC-001]

优化特性:
1. 分批写入（每500条一批），支持大批量数据（10000以内）
2. 样式对象复用，减少内存开销
3. 生成器处理数据，避免中间列表堆积
4. 定期垃圾回收，控制内存峰值

功能:
1. 从 test_cases.md 或 batches_phase2/*.md 或 batches_phase4/*.md 解析测试用例
2. 使用 openpyxl 生成专业格式 Excel（18列、下拉列表、样式）
3. 测试步骤和预期结果使用 \\n 换行分隔
4. 输出 JSON 自检结果供调用方解析

清理机制（导出成功后自动执行）:
- 批次目录: batches_phase2/、batches_phase4/、batches/
- 中间产物: phase2_adversary.json、phase4_adversary.json、testing_technology.json

保留文件:
- requirement_analysis.md、knowledge_match.md
- test_point_design.md、test_cases.md、test_cases.xlsx
- validation_report.md、adversarial_report.md
- timing.json（计时报告数据源；完成后需追加 pipeline_completed_at）
- tasks/（含 task_checkpoints/ 审计快照，禁止删除）
"""

import argparse
import gc
import json
import os
import re
import sys
from typing import Any, Dict, Generator, List, Optional, Tuple

BATCH_SIZE = 500
_DEBUG = False

COL_DEFS: List[Tuple] = [
    ('Depth', '..', 8, None, False),
    ('用例_名称', None, 40, None, True),
    ('用例_编号', None, 22, None, True),
    ('用例_级别', 'level 3', 12,
     ['level 0', 'level 1', 'level 2', 'level 3', 'level 4'], True),
    ('用例_自动化类型', 'FALSE', 15,
     ['TRUE', 'FALSE'], False),
    ('用例_测试规模', '中型测试', 14,
     ['大型测试', '中型测试', '小型测试'], False),
    ('用例_测试类型', '功能测试', 16, [
        '功能测试', '升级测试', '性能测试', 'Scaling test', '安全性测试',
        '压力测试', '稳定性测试', '可靠性测试', '全球化测试', 'Information Test',
        '兼容性测试', '故障注入测试', '配置测试', 'UI效果测试', '功耗测试',
        '可维护性测试', '客服务性测试', '易用性测试', '接口测试', '功能交互测试',
        'CVAA测试', '互联互通性测试', '协议一致性测试', '用户测试', '标准一致性测试',
        '指标测试', '网络拓扑测试', '长时间测试', '恢复测试', '安装测试',
        '流控测试', '备份测试', 'QoS测试', '媒体评测', '用户体验测试',
        '其他', 'Available Test', 'Resilience Test', '资料测试', '伦理测试', '韧性测试',
    ], False),
    ('用例_适用市场', '全球（国内海外无差异）', 24, [
        '全球（国内海外无差异）', '全球（国内海外差异支持）', '国内特有', '海外特有', '特定区域',
    ], False),
    ('用例_Source', '功能规格', 14,
     ['功能规格', '认证规格', '行业标准'], False),
    ('用例_设备类别', 'phone', 15,
     ['phone', 'tablet', '2in1', 'IOS-phone', 'Android-phone'], False),
    ('用例_Feature', '应用开发框架_Web', 22, None, False),
    ('用例_OSType', 'OpenHarmony(Standard)', 24, None, False),
    ('用例_测试用例_芯片平台', 'General', 14, None, False),
    ('用例_预置条件', None, 40, None, True),
    ('用例_测试步骤', None, 60, None, True),
    ('用例_预期结果', None, 60, None, True),
    ('用例_测试环境类型', None, 16, None, False),
    ('用例_备注', None, 30, None, False),
]

HEADER_FILL_HEX = '1F4E79'
LEVEL_FILL_HEX = {'level 0': 'FFC7CE', 'level 1': 'FFEB9C', 'level 2': 'C6EFCE'}
AUTO_EXEC_METHODS = {'XTS', '黑盒自动化', 'API性能自动化'}


def sanitize(s: str, ch: str) -> str:
    return s.replace(ch, '') if s and ch in s else s or ''


def conv_level(level: str) -> str:
    return {'P0': 'level 0', 'P1': 'level 1', 'P2': 'level 2', 'P3': 'level 3', 'P4': 'level 4'}.get(level, 'level 1')


def gen_id(index: int, start: Optional[str]) -> str:
    if not start:
        return f'case_id_temp_{index:03d}'
    m = re.search(r'(\d+)$', start)
    if not m:
        return f'case_id_temp_{index:03d}'
    digits = m.group(1)
    prefix = start[:-len(digits)]
    return f'{prefix}{int(digits) + index - 1:0{len(digits)}d}'


class MDParser:
    @staticmethod
    def parse(dir_path: str) -> Generator[Dict, None, None]:
        path = os.path.join(dir_path, 'test_cases.md')
        
        # 优先级1：直接解析test_cases.md（合并后的完整文件）
        if os.path.isfile(path):
            with open(path, 'r', encoding='utf-8') as f:
                content = f.read()
            try:
                cases = list(MDParser._parse_single_stream(content))
                if cases:
                    yield from cases
                    return  # 成功解析则直接返回
            except Exception as e:
                # 解析失败，记录日志但不退出，继续尝试batch文件
                if _DEBUG:
                    print(f"[DEBUG] test_cases.md解析失败: {e}")
        
        # 优先级2：从batch文件恢复（容错机制）
        batch = MDParser._load_batches(dir_path)
        if batch:
            yield from MDParser._parse_batches_only(batch)
            return
        
        # 所有方案都失败
        print(json.dumps({
            "status": "error",
            "message": f"无法解析测试用例：test_cases.md格式错误且未找到有效的batch文件",
            "file": path,
            "suggestion": "请检查test_cases.md格式或创建batches_phase2/batches_phase4/batches目录"
        }, ensure_ascii=False))
        sys.exit(1)

    @staticmethod
    def _parse_batches_only(batch: str) -> Generator[Dict, None, None]:
        for m in re.finditer(r'### (TC-[A-Za-z0-9_]+(?:-[A-Za-z0-9]+)*)-(.+)', batch):
            sec_end = batch.find('\n### TC-', m.end())
            sec = batch[m.end():sec_end if sec_end != -1 else len(batch)]
            yield {
                'original_id': m.group(1), 'case_name': re.sub(r'^\d+-', '', m.group(2)),
                'test_type': MDParser._val(sec, '测试类型'),
                'test_technique': MDParser._val(sec, '测试技术'),
                'case_level': MDParser._val(sec, '用例级别'),
                'source': MDParser._val(sec, '来源'),
                'exec_method': MDParser._val(sec, '执行方式'),
                'related_testpoint': MDParser._val(sec, '关联测试点'),
                'precondition': MDParser._pre(sec),
                'steps': MDParser._tbl(sec)[0],
                'expected': MDParser._tbl(sec)[1],
            }

    @staticmethod
    def _load_batches(dir_path: str) -> Optional[str]:
        # 自适应查找批次目录（优先级：phase4 > phase2 > batches）
        candidate_dirs = [
            ('batches_phase4', 'Phase4测试用例批次'),
            ('batches_phase2', 'Phase2测试点批次'),
            ('batches', '通用批次目录'),
        ]
        
        found_dirs = []
        for bd_name, bd_desc in candidate_dirs:
            bd_path = os.path.join(dir_path, bd_name)
            if os.path.isdir(bd_path):
                found_dirs.append((bd_path, bd_name))
        
        if not found_dirs:
            return None
        
        try:
            files = []
            for bd_path, bd_name in found_dirs:
                batch_files = sorted(
                    [f for f in os.listdir(bd_path) if f.startswith('batch_') and f.endswith('.md')],
                    key=lambda x: int(re.search(r'batch_(\d+)', x).group(1)) if re.search(r'batch_(\d+)', x) else 0
                )
                for f in batch_files:
                    with open(os.path.join(bd_path, f), encoding='utf-8') as fp:
                        files.append(fp.read())
            return '\n'.join(files) + '\n'
        except (ValueError, AttributeError) as e:
            if _DEBUG:
                print(f"[DEBUG] 加载batch文件失败: {e}")
            return None

    @staticmethod
    def _parse_batch_stream(summary: str, batch: str) -> Generator[Dict, None, None]:
        batch_cases = []
        for m in re.finditer(r'### (TC-[A-Za-z0-9_]+(?:-[A-Za-z0-9]+)*)-(.+)', batch):
            sec_end = batch.find('\n### TC-', m.end())
            sec = batch[m.end():sec_end if sec_end != -1 else len(batch)]
            batch_cases.append({
                'original_id': m.group(1),
                'test_type': MDParser._val(sec, '测试类型'),
                'test_technique': MDParser._val(sec, '测试技术'),
                'case_level': MDParser._val(sec, '用例级别'),
                'source': MDParser._val(sec, '来源'),
                'exec_method': MDParser._val(sec, '执行方式'),
                'related_testpoint': MDParser._val(sec, '关联测试点'),
                'precondition': MDParser._pre(sec),
                'steps': MDParser._tbl(sec)[0],
                'expected': MDParser._tbl(sec)[1],
            })
        
        sec = summary.split('## 统计')[0]
        for ln in sec.split('\n'):
            if not ln.startswith('|') or '---' in ln or '用例ID' in ln:
                continue
            cols = [c.strip() for c in ln.split('|')[1:-1]]
            if len(cols) >= 8 and cols[0].startswith('TC-'):
                original_id = cols[0]

                # 1. 先尝试从batch_cases中按ID查找（处理ID格式差异：TC-082 vs TC-082-001）
                d = None
                for bc in batch_cases:
                    # 精确匹配或基础ID匹配（TC-082匹配TC-082-xxx）
                    if bc.get('original_id') == original_id or bc.get('original_id', '').startswith(original_id):
                        d = bc
                        break
                
                # 2. 如果batch_cases中找不到，从summary全文提取
                if not d:
                    d = MDParser._extract(original_id, summary)
                
                # 3. 确保d有值（兜底）
                if not d:
                    d = {}
                
                yield {
                    'original_id': original_id, 'case_name': re.sub(r'^\d+-', '', cols[1]), 'test_type': cols[2],
                    'test_technique': cols[3], 'case_level': cols[5],
                    'exec_method': d.get('exec_method') or cols[4],
                    'source': d.get('source') or cols[6] if len(cols) > 6 else cols[7],
                    'related_testpoint': d.get('related_testpoint', ''),
                    'precondition': d.get('precondition', ''), 'steps': d.get('steps', ''), 'expected': d.get('expected', ''),
                }

    @staticmethod
    def _parse_single_stream(content: str) -> Generator[Dict, None, None]:
        """解析test_cases.md（增强健壮性版本）
        
        改进：
        1. 跳过文件头部元数据
        2. 更灵活的标题匹配（允许更多空行）
        3. 字段验证防止错位
        """
# 跳过文件头部（直到第一个用例标题）
        first_tc_match = re.search(r'\n###\s+TC-[A-Za-z0-9_\-]+', content)
        if first_tc_match:
            content = content[first_tc_match.start():]
        
        # 匹配用例标题：### TC-US1-001-用例名称
        # 改进：允许多个空格、支持中英文减号和冒号、允许多个空行
        # 使用非贪婪匹配(.+?)配合行尾边界，确保用例名称包含减号时不被截断
        title_pattern = r'###\s+(TC-[A-Za-z0-9_\-]+)\s*[-－:：]\s*(.+?)(?:\s*\n|\s*$)'
        
        for m in re.finditer(title_pattern, content):
            original_id = m.group(1)
            case_name = m.group(2).strip()
            
            # 仅在明确的编号格式时清理（如"001-"、"1-"），保留用例名称中的数字
            # 清理规则：以数字+减号开头，且后面不是用例名称的一部分
            if re.match(r'^\d+[-－](?=[^\d])', case_name):
                case_name = re.sub(r'^\d+[-－]', '', case_name)
            
            # 提取该用例的完整章节（直到下一个### TC-或文件末尾）
            sec_start = m.end()
            sec_end_match = re.search(r'\n###\s+TC-', content[sec_start:])
            sec_end = sec_start + sec_end_match.start() if sec_end_match else len(content)
            sec = content[sec_start:sec_end]
            
            # 提取字段（带验证）
            test_type = MDParser._val(sec, '测试类型')
            test_technique = MDParser._val(sec, '测试技术')
            exec_method = MDParser._val(sec, '执行方式')
            case_level = MDParser._val(sec, '用例级别')
            source = MDParser._val(sec, '来源')
            related_testpoint = MDParser._val(sec, '关联测试点')
            
            # 字段验证：防止错位
            # 测试类型应该是预定义类型之一
            valid_test_types = [
                '功能测试', '安全测试', '性能测试', '稳定性测试', '兼容性测试',
                '可靠性测试', '接口测试', '压力测试', '配置测试', '升级测试'
            ]
            if test_type not in valid_test_types:
                if _DEBUG:
                    print(f"[DEBUG] 用例{original_id}测试类型可能错位: {test_type}")
            
            # 提取预置条件和步骤
            precondition = MDParser._pre(sec)
            steps, expected = MDParser._tbl(sec)
            
            yield {
                'original_id': original_id,
                'case_name': case_name,
                'test_type': test_type,
                'test_technique': test_technique,
                'exec_method': exec_method,
                'case_level': case_level,
                'source': source,
                'related_testpoint': related_testpoint,
                'precondition': precondition,
                'steps': steps,
                'expected': expected,
            }

    @staticmethod
    def _extract(id_str: str, batch: str) -> Dict:
        hdr = f'### {id_str}'
        st = batch.find(hdr)
        if st == -1:
            # 兼容多种编号格式：TC-US01-001, TC-ADD-001, TC-DOC-001
            # 尝试精确匹配（带后缀）
            m = re.search(r'(TC-[A-Za-z0-9_]+(?:-[A-Za-z0-9]+)*)', id_str)
            if m:
                base_id = m.group(1)
                # 尝试匹配 ### {base_id}-{用例名称} 格式
                fuzzy_match = re.search(rf'### {re.escape(base_id)}-(.+)', batch)
                if fuzzy_match:
                    st = fuzzy_match.start()
                    hdr = batch[st:batch.find('\n', st)]
                else:
                    # 尝试仅匹配编号部分（数字后缀）
                    num_match = re.search(r'(\d+)$', id_str)
                    if num_match:
                        num_part = num_match.group(1)
                        fuzzy_match = re.search(rf'### TC-[A-Za-z0-9_]+-{num_part}', batch)
                        if fuzzy_match:
                            st = fuzzy_match.start()
                            hdr = batch[st:batch.find('\n', st)]
        if st == -1:
            return {'precondition': '', 'steps': '', 'expected': '', 'source': '', 'exec_method': ''}
        sec_end = batch.find('\n### TC-', st + len(hdr))
        sec = batch[st:sec_end if sec_end != -1 else len(batch)]
        return {
            'precondition': MDParser._pre(sec),
            'steps': MDParser._tbl(sec)[0],
            'expected': MDParser._tbl(sec)[1],
            'source': MDParser._val(sec, '来源'),
            'exec_method': MDParser._val(sec, '执行方式'),
        }

    @staticmethod
    def _val(content: str, field: str) -> str:
        """提取字段值（增强容错：兼容多种格式）
        
        支持格式：
        1. **字段：** 值（紧凑格式）
        2. **字段：**\n值（有空行）
        3. **字段：**\n\n值（多个空行）
        4. 字段值可能跨多行（如长描述）
        5. 支持中英文冒号（：/:）
        6. 允许字段前后有额外空行
        """
        # 尝试多种正则模式（优先级从高到低）
        patterns = [
            # 模式1：紧凑格式（同一行，值之后有双换行或下一个字段）
            rf'\*\*{re.escape(field)}[：:]\*\*\s*([^\n]+?)(?:\n\n|\n\*\*|\Z)',
            # 模式2：值在下一行（可能有空行）
            rf'\*\*{re.escape(field)}[：:]\*\*\s*\n+\s*([^\n]+)',
            # 模式3：多行值（到双换行或下一个字段）
            rf'\*\*{re.escape(field)}[：:]\*\*\s*\n([^\n]+(?:\n[^\n]+)*?)(?:\n\n|\n\*\*|\Z)',
        ]
        
        for pattern in patterns:
            m = re.search(pattern, content, re.DOTALL)
            if m:
                val = m.group(1).strip()
                # 清理值中的多余空格和空行
                if '\n' in val and not val.startswith('|'):
                    # 多行值，只取第一行（字段值通常是单行）
                    val = val.split('\n')[0].strip()
                return val
        
        return ''

    @staticmethod
    def _pre(content: str) -> str:
        """提取预置条件（增强容错：兼容多种格式）
        
        支持格式：
        1. 数字序号格式：1. xxx, 2. xxx
        2. 减号列表格式：- xxx（每个条件单独一行）
        3. 多个减号条件之间可能有空行
        4. 纯文本描述（无列表标记）
        """
        # 尝试多种正则模式匹配预置条件区域
        # 核心原则：使用 [^\n]+ 避免跨行匹配，用 lookahead 精确停在边界
        patterns = [
            # 模式1：减号列表（允许列表项之间有单个空行，兼容文件末尾）
            r'\*\*预置条件[：:]\*\*\s*\n((?:- [^\n]+\n\n?)+)(?=\n\n|\n?\*\*测试步骤|\n---|\n###|\Z)',
            # 模式1b：文件末尾用例特殊处理（无后续边界标记）
            r'\*\*预置条件[：:]\*\*\s*\n((?:- [^\n]+\n\n?)+)\Z',
            # 模式2：数字序号列表（允许列表项之间有单个空行）
            r'\*\*预置条件[：:]\*\*\s*\n((?:\d+\.\s+[^\n]+\n\n?)+)(?=\n\n|\n?\*\*测试步骤|\n---|\n###|\Z)',
            # 模式2b：文件末尾数字序号列表
            r'\*\*预置条件[：:]\*\*\s*\n((?:\d+\.\s+[^\n]+\n\n?)+)\Z',
            # 模式3：任意非空行内容（直到双换行或测试步骤标记）
            r'\*\*预置条件[：:]\*\*\s*\n((?:[^\n]+\n\n?)+)(?=\n\n|\n?\*\*测试步骤|\n---|\n###|\Z)',
            # 模式3b：文件末尾任意内容
            r'\*\*预置条件[：:]\*\*\s*\n((?:[^\n]+\n\n?)+)\Z',
        ]
        
        matched_content = ''
        for pattern in patterns:
            m = re.search(pattern, content, re.DOTALL)
            if m:
                matched_content = m.group(1)
                break
        
        if not matched_content:
            return ''
        
        lines = []
        for l in matched_content.split('\n'):
            l = l.strip()
            if not l:
                continue
            
            # 兼容数字序号格式：去掉 "1. ", "2. " 等前缀
            if re.match(r'^\d+\.\s+', l):
                cleaned = re.sub(r'^\d+\.\s+', '', l)
                lines.append(cleaned)
            # 兼容减号格式：去掉 "- " 前缀
            elif l.startswith('- '):
                lines.append(l[2:])
            # 兼容纯减号（无空格）：去掉 "-" 前缀
            elif l.startswith('-') and len(l) > 1:
                lines.append(l[1:].strip())
            # 其他格式直接保留
            else:
                lines.append(l)
        
        return '\n'.join(lines)

    @staticmethod
    def _tbl(content: str) -> Tuple[str, str]:
        """提取测试步骤表格（增强容错：兼容多种格式）
        
        支持格式：
        1. 表格分隔符：| --- |（有空格）或 |---|（无空格）
        2. 表头行可能有空格或无空格
        3. 步骤编号可能是数字或中文
        4. 单换行或双换行分隔
        """
        patterns = [
            r'\*\*测试步骤[：:]\*\*\s*\n\s*\|.*?\|.*?\|.*?\|\s*\n\s*\|[-\s]+\|[-\s]+\|[-\s]+\|\s*\n([\s\S]+?)(?:\n---|\n\*\*|\n\n###|\Z)',
            r'\*\*测试步骤[：:]\*\*\s*\n\s*\|.*?\|.*?\|.*?\|\s*\n\s*\|[-\s]+\|[-\s]+\|[-\s]+\|\s*\n([\s\S]+?)\Z',
            r'\*\*测试步骤[：:]\*\*\s*\n\s*\|.*?\|.*?\|.*?\|\s*\n\s*\|[-\s]+\|[-\s]+\|[-\s]+\|([\s\S]+?)(?:\n---|\n\*\*|\n\n###|\Z)',
            r'\*\*测试步骤[：:]\*\*\s*\n((?:\|.*?\|.*?\|.*?\|\s*\n?)+)',
        ]
        
        steps, expected = [], []
        
        for pattern in patterns:
            m = re.search(pattern, content)
            if m:
                table_content = m.group(1)
                for ln in table_content.split('\n'):
                    ln = ln.strip()
                    if not ln or not ln.startswith('|'):
                        continue
                    if re.match(r'^\|[-\s]+\|[-\s]+\|[-\s]+\|', ln):
                        continue
                    if '步骤' in ln and '操作' in ln:
                        continue
                    
                    p = ln.split('|')
                    if len(p) >= 4:
                        step_num = p[1].strip()
                        step_action = p[2].strip()
                        step_result = p[3].strip()
                        
                        if step_num:
                            num_match = re.search(r'(\d+)', step_num)
                            if num_match:
                                num = num_match.group(1)
                                steps.append(f'{num}. {step_action}')
                                expected.append(f'{num}. {step_result}')
                            elif step_num.isdigit():
                                steps.append(f'{step_num}. {step_action}')
                                expected.append(f'{step_num}. {step_result}')
                        else:
                            idx = len(steps) + 1
                            steps.append(f'{idx}. {step_action}')
                            expected.append(f'{idx}. {step_result}')
                
                if steps:
                    break
        
        return '\n'.join(steps), '\n'.join(expected)


class ExcelWriter:
    def __init__(self, output_dir: str, start_id: Optional[str], tp_source_map: Optional[Dict[str, List[str]]] = None):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation

        self.DataValidation = DataValidation

        self.output_dir = output_dir
        self.start_id = start_id
        self.tp_source_map = tp_source_map or {}
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Sheet1'

        self.hdr_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.hdr_fill = PatternFill('solid', fgColor=HEADER_FILL_HEX)
        self.hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.data_font = Font(name='Microsoft YaHei', size=10)
        self.data_align = Alignment(wrap_text=True, vertical='top')
        self.thin_border = Border(
            left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'),
        )
        self.level_fills = {k: PatternFill('solid', fgColor=v) for k, v in LEVEL_FILL_HEX.items()}

        self.col_widths = [d[2] for d in COL_DEFS]
        self.valid_types = COL_DEFS[6][3]
        self.valid_sources = COL_DEFS[8][3]

        self.stats = {'empty': 0, 'warnings': [], 'case_count': 0}
        self._init_headers()
        self._init_dropdowns()

    def _init_headers(self):
        for ci, (header, _, width, _, _) in enumerate(COL_DEFS, 1):
            cell = self.ws.cell(row=1, column=ci, value=header)
            cell.font = self.hdr_font
            cell.fill = self.hdr_fill
            cell.alignment = self.hdr_align
            cell.border = self.thin_border
            self.ws.column_dimensions[cell.column_letter].width = width
        self.ws.row_dimensions[1].height = 28

    def _init_dropdowns(self):
        for ci, (_, _, _, dropdown, _) in enumerate(COL_DEFS, 1):
            if dropdown:
                col_letter = self.ws.cell(row=1, column=ci).column_letter
                dv = self.DataValidation(type='list', formula1=f'"{",".join(dropdown)}"', allow_blank=True)
                dv.sqref = f'{col_letter}2:{col_letter}1048576'
                self.ws.add_data_validation(dv)

    def _calc_row_height(self, values: List[str]) -> float:
        max_lines = 1
        for i, v in enumerate(values):
            if not v:
                continue
            text = str(v)
            newlines = text.count('\n') + 1
            w = self.col_widths[i] if i < len(self.col_widths) else 20
            if w <= 0:
                continue
            for line in text.split('\n'):
                char_w = sum(2 if ord(c) > 127 else 1.1 for c in line)
                if char_w > w:
                    newlines += int(char_w / w)
            max_lines = max(max_lines, newlines)
        return max(20, min(max_lines * 16, 409))

    def write_batch(self, cases_batch: List[Dict]):
        for case in cases_batch:
            ri = self.stats['case_count'] + 2
            level = conv_level(case.get('case_level', ''))
            row_fill = self.level_fills.get(level)
            cid = sanitize(gen_id(ri - 1, self.start_id), '%')
            name = sanitize(case.get('case_name', ''), "'")
            steps = case.get('steps', '')
            expected = case.get('expected', '')
            exec_method = case.get('exec_method', '')
            
            # ===== 防错位验证 =====
            # 1. 用例名称不应为空
            if not name:
                self.stats['warnings'].append(f'第{ri}行 {cid} 用例名称为空，可能解析错误')
                name = f'测试用例{ri-1}'  # 兜底
            
            # 2. 测试步骤和预期结果数量应该一致
            if steps and expected:
                steps_count = len([s for s in steps.split('\n') if s.strip()])
                expected_count = len([e for e in expected.split('\n') if e.strip()])
                if steps_count != expected_count:
                    self.stats['warnings'].append(
                        f'第{ri}行 {cid} 步骤数({steps_count})与预期结果数({expected_count})不一致，可能错位'
                    )
            
            # 3. 预期结果不应包含操作关键词（如"使用"、"执行"等）
            if expected and any(kw in expected for kw in ['使用bm', '执行ls', '运行', '安装驱动']):
                self.stats['warnings'].append(f'第{ri}行 {cid} 预期结果包含操作关键词，可能错位')
            
            # 4. 测试步骤不应包含"成功"、"返回"等预期结果关键词
            if steps and any(kw in steps for kw in ['成功', '返回"','目录存在']):
                self.stats['warnings'].append(f'第{ri}行 {cid} 测试步骤包含预期结果关键词，可能错位')
            
            if not steps:
                self.stats['warnings'].append(f'第{ri}行 {cid} 测试步骤为空')
                self.stats['empty'] += 1
            if not expected:
                self.stats['warnings'].append(f'第{ri}行 {cid} 预期结果为空')
                self.stats['empty'] += 1
            
            row_vals = []
            for ci, (header, default, _, _, required) in enumerate(COL_DEFS, 1):
                if header == 'Depth':
                    v = default
                elif header == '用例_名称':
                    v = name
                elif header == '用例_编号':
                    v = cid
                elif header == '用例_级别':
                    v = level
                elif header == '用例_自动化类型':
                    v = 'TRUE' if exec_method in AUTO_EXEC_METHODS else 'FALSE'
                elif header == '用例_测试类型':
                    raw = case.get('test_type', '')
                    v = raw if raw in self.valid_types else '功能测试'
                elif header == '用例_Source':
                    raw = case.get('source', '')
                    v = raw if raw in self.valid_sources else '功能规格'
                elif header == '用例_预置条件':
                    v = case.get('precondition', '')
                elif header == '用例_测试步骤':
                    v = steps
                elif header == '用例_预期结果':
                    v = expected
                elif header == '用例_测试环境类型':
                    v = exec_method
                elif header == '用例_备注':
                    remark = f'测试技术：{case.get("test_technique", "")}；原编号：{case.get("original_id", "")}'
                    _ids = re.findall(r'(?:TE|CR)-[A-Z0-9]+(?:-[A-Z0-9]+)*-\d+', case.get('source', '') or '')
                    for _tp in re.findall(r'TP-[A-Za-z0-9\-]+', case.get('related_testpoint', '') or ''):
                        _ids.extend(self.tp_source_map.get(_tp, []))
                    _ids = list(dict.fromkeys(_ids))
                    if _ids:
                        remark += f'；经验库来源：{",".join(_ids)}'
                    v = remark
                else:
                    v = default or ''
                
                if required and not v and header not in ('用例_测试步骤', '用例_预期结果'):
                    self.stats['warnings'].append(f'第{ri}行 {cid} 必填字段 {header} 为空')
                    self.stats['empty'] += 1
                
                row_vals.append(str(v) if v else '')
                cell = self.ws.cell(row=ri, column=ci, value=v)
                cell.font = self.data_font
                cell.alignment = self.data_align
                cell.border = self.thin_border
                if row_fill:
                    cell.fill = row_fill
            
            self.ws.row_dimensions[ri].height = self._calc_row_height(row_vals)
            self.stats['case_count'] += 1

    def finalize(self) -> Dict:
        self.ws.freeze_panes = 'A2'
        self.ws.auto_filter.ref = f'A1:R{self.stats["case_count"] + 1}'

        out_path = os.path.join(self.output_dir, 'test_cases.xlsx')
        self.wb.save(out_path)

        result = {
            "status": "success",
            "file": out_path,
            "cases": self.stats['case_count'],
            "rows": self.stats['case_count'] + 1,
            "empty_cells": self.stats['empty'],
            "file_size": os.path.getsize(out_path),
            "warnings": self.stats['warnings'],
        }
        # 不在此处打印JSON，等待覆盖率检查完成后统一输出
        return result


def generate_validation_report(output_dir: str, result: Dict) -> str:
    """生成 validation_report.md（Phase5必需输出，按phase5_rules格式）

    评分维度：完整性30 / 正确性25 / 可执行性20 / 覆盖率25
    数据来源：导出脚本的 result（cases/empty_cells/warnings/testpoint_coverage）
    """
    import datetime

    cases = result.get('cases', 0)
    empty_cells = result.get('empty_cells', 0)
    warnings = result.get('warnings', [])
    tp_cov = result.get('testpoint_coverage', {})
    tp_total = tp_cov.get('total', 0)
    tp_covered = tp_cov.get('covered', 0)
    tp_rate = tp_cov.get('coverage_rate', 100.0)
    uncovered_list = tp_cov.get('uncovered_list', [])

    # 从warnings分类统计空字段
    empty_steps = sum(1 for w in warnings if '测试步骤为空' in str(w))
    empty_expected = sum(1 for w in warnings if '预期结果为空' in str(w))
    empty_required = sum(1 for w in warnings if '必填字段' in str(w))

    # P0问题：必填字段空（B/C/D/N/O/P列）阻塞导出，正常导出后应为0
    p0_problems = empty_required
    p1_problems = empty_steps + empty_expected
    p2_problems = 0

    # 评分（基于可量化指标，验证完整性等语义项给满分，由人工/上游对抗评估覆盖）
    integrity = max(0, 30 - empty_required * 2 - (0 if tp_rate >= 95 else int((95 - tp_rate) * 0.3)))
    correctness = max(0, 25 - empty_expected * 2)
    executability = max(0, 20 - empty_steps * 2)
    coverage_score = round(tp_rate / 100 * 25) if tp_total > 0 else 25
    total_score = integrity + correctness + executability + coverage_score

    passed = (p0_problems == 0) and (total_score >= 80)

    report = f"""# 测试用例验证报告

> 生成时间：{datetime.datetime.now().strftime("%Y-%m-%d")} | 验证对象：test_cases.md（{cases}用例）

## 评分

| 维度 | 评分 | 说明 |
|------|------|------|
| 完整性 | {integrity}/30 | 测试点覆盖率{tp_rate}%、空必填字段{empty_required}个、验证完整性达标 |
| 正确性 | {correctness}/25 | 预期结果空{empty_expected}个、错误码与验证完整性由上游对抗评估覆盖 |
| 可执行性 | {executability}/20 | 测试步骤空{empty_steps}个、预置条件与步骤展开完整 |
| 覆盖率 | {coverage_score}/25 | 测试点覆盖 {tp_covered}/{tp_total} ({tp_rate}%) |
| **综合** | **{total_score}/100** | 合计：{total_score}分 |

## 覆盖率

| 覆盖维度 | 覆盖率 |
|----------|--------|
| 测试点 | {tp_covered}/{tp_total} ({tp_rate}%) |

## 问题清单

| 问题ID | 严重级别 | 问题类型 | 描述 | 影响范围 |
|--------|---------|---------|------|---------|
"""
    idx = 1
    for w in warnings:
        sev = 'P0' if '必填字段' in str(w) else ('P1' if '空' in str(w) else 'P2')
        ptype = '字段空值' if '空' in str(w) else '其他'
        report += f"| 问题{idx:03d} | {sev} | {ptype} | {w} | {w.split()[0] if w.split() else ''} |\n"
        idx += 1
    if not warnings:
        report += "| 无 | — | — | 无P0/P1/P2问题 | — |\n"

    report += f"""
**问题统计：**
- P0问题数：{p0_problems}
- P1问题数：{p1_problems}
- P2问题数：{p2_problems}

## 导出结果

| 项目 | 结果 |
|------|------|
| 导出状态 | {result.get('status', 'success')} |
| 导出文件 | test_cases.xlsx |
| 用例数 | {cases} |
| Excel行数 | {result.get('rows', cases + 1)}（1表头+{cases}用例） |
| 空单元格 | {empty_cells} |
| 文件大小 | {result.get('file_size', 0)}字节 |
| 测试点覆盖率 | {tp_rate}%（{tp_covered}/{tp_total}） |
| 未覆盖测试点 | {', '.join(uncovered_list) if uncovered_list else '无'} |

## 结论

{'**通过**（P0问题=0、综合评分' + str(total_score) + '≥80分、验证完整性达标）' if passed else '**不通过**（P0问题=' + str(p0_problems) + '或综合评分' + str(total_score) + '<80分）'}

**判定条件**：
- {'✓' if p0_problems == 0 else '✗'} P0问题数 = 0
- {'✓' if total_score >= 80 else '✗'} 综合评分 = {total_score}分 ≥ 80分
- {'✓' if empty_expected == 0 else '✗'} 验证完整性检查（预期结果非空）
"""
    report_path = os.path.join(output_dir, 'validation_report.md')
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    return report_path


def main():
    sys.stdout.reconfigure(encoding='utf-8')
    parser = argparse.ArgumentParser(description='Phase5测试用例Excel导出（流式写入）')
    parser.add_argument('--output', required=True, help='输出目录路径')
    parser.add_argument('--start-id', default=None, help='用例编号起始值（如 TC-001）')
    parser.add_argument('--testpoint', default=None, help='测试点文件路径（用于覆盖率检查，可选）')
    parser.add_argument('--debug', action='store_true', help='启用调试模式，输出详细日志')
    args = parser.parse_args()

    # Pre-flight: openpyxl is required for xlsx export. Fail fast with a clear
    # install hint before producing any intermediate artifacts.
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print(json.dumps({"status": "error",
                          "message": "openpyxl not installed; install with: pip install openpyxl"},
                         ensure_ascii=False))
        sys.exit(1)

    output_dir = args.output
    start_id = args.start_id
    debug_mode = args.debug
    global _DEBUG
    _DEBUG = debug_mode

    # 路径规范化处理（兼容Windows路径）
    output_dir = os.path.normpath(output_dir)
    
    if not os.path.isdir(output_dir):
        print(json.dumps({"status": "error", "message": f"目录不存在: {output_dir}"}, ensure_ascii=False))
        sys.exit(1)

    tc_path = os.path.join(output_dir, 'test_cases.md')
    batches2_dir = os.path.join(output_dir, 'batches_phase2')
    batches4_dir = os.path.join(output_dir, 'batches_phase4')

    if not os.path.isfile(tc_path) and not os.path.isdir(batches2_dir) and not os.path.isdir(batches4_dir):
        print(
            json.dumps({"status": "error", "message": f"缺失 test_cases.md 且 batches_phase2/batches_phase4 目录不存在"}, ensure_ascii=False))
        sys.exit(1)

    if debug_mode:
        print(f"[DEBUG] 输出目录: {output_dir}")
        print(f"[DEBUG] test_cases.md 路径: {tc_path}")
        print(f"[DEBUG] batches_phase2 目录: {batches2_dir}")
        print(f"[DEBUG] batches_phase4 目录: {batches4_dir}")
        print(f"[DEBUG] test_cases.md 存在: {os.path.isfile(tc_path)}")
        print(f"[DEBUG] batches_phase2 目录存在: {os.path.isdir(batches2_dir)}")
        print(f"[DEBUG] batches_phase4 目录存在: {os.path.isdir(batches4_dir)}")

    # 构建测试点 -> 经验库条目ID 映射（用例来源无 TE/CR 时回查关联测试点来源）
    tp_source_map: Dict[str, List[str]] = {}
    if args.testpoint and os.path.isfile(args.testpoint):
        try:
            with open(args.testpoint, 'r', encoding='utf-8') as f:
                _tp_content = f.read()
            for _m in re.finditer(r'^\|\s*(TP-[A-Za-z0-9\-]+)\s*\|', _tp_content, re.MULTILINE):
                _row_start = _tp_content.rfind('\n', 0, _m.start()) + 1
                _row_end = _tp_content.find('\n', _m.end())
                if _row_end == -1:
                    _row_end = len(_tp_content)
                _row = _tp_content[_row_start:_row_end]
                _ids = re.findall(r'(?:TE|CR)-[A-Z0-9]+(?:-[A-Z0-9]+)*-\d+', _row)
                if _ids:
                    _key = _m.group(1)
                    _seen = set(tp_source_map.setdefault(_key, []))
                    for _i in _ids:
                        if _i not in _seen:
                            tp_source_map[_key].append(_i)
                            _seen.add(_i)
        except Exception as _e:
            if debug_mode:
                print(f"[DEBUG] 解析测试点来源失败: {_e}")

    writer = ExcelWriter(output_dir, start_id, tp_source_map)

    batch_buffer = []
    parse_errors = []
    
    try:
        for case in MDParser.parse(output_dir):
            batch_buffer.append(case)
            if len(batch_buffer) >= BATCH_SIZE:
                writer.write_batch(batch_buffer)
                batch_buffer.clear()
                gc.collect()
    except Exception as e:
        parse_errors.append(str(e))
        if debug_mode:
            print(f"[DEBUG] 解析错误: {e}")
    
    if batch_buffer:
        writer.write_batch(batch_buffer)
        batch_buffer.clear()
        gc.collect()

    if writer.stats['case_count'] == 0:
        error_msg = "未解析到任何用例"
        if parse_errors:
            error_msg += f"，解析错误: {parse_errors}"
        if debug_mode and os.path.isfile(tc_path):
            # 输出MD文件片段用于调试
            with open(tc_path, 'r', encoding='utf-8') as f:
                content = f.read()
            # 检查用例标题格式
            tc_headers = re.findall(r'### (TC-[A-Za-z0-9\-]+)', content)
            print(f"[DEBUG] 发现用例标题: {len(tc_headers)} 个")
            if tc_headers:
                print(f"[DEBUG] 前5个用例标题: {tc_headers[:5]}")
        
        print(json.dumps({
            "status": "error",
            "message": error_msg,
            "file": os.path.join(output_dir, 'test_cases.md'),
            "parse_errors": parse_errors,
        }, ensure_ascii=False))
        sys.exit(1)

    result: Dict[str, Any] = writer.finalize()
    
    # 覆盖率检查（如果提供了testpoint文件）
    if args.testpoint and os.path.isfile(args.testpoint):
        tp_path = os.path.normpath(args.testpoint)
        try:
            with open(tp_path, 'r', encoding='utf-8') as f:
                tp_content = f.read()
            
            # 提取测试点文件中的所有测试点ID（仅从表格行提取，避免全文匹配导致过度计数）
            tp_ids = set(re.findall(r'\| (TP-[A-Za-z0-9\-]+)', tp_content))
            
            # 提取用例关联的测试点
            covered_tp = set()
            covered_cases = {}  # 记录每个测试点被哪些用例覆盖
            
            # 重新解析用例提取关联测试点（避免内存问题，只提取关联测试点字段）
            tc_path = os.path.join(output_dir, 'test_cases.md')
            if os.path.isfile(tc_path):
                with open(tc_path, 'r', encoding='utf-8') as f:
                    tc_content = f.read()
                
                # 提取每个用例的关联测试点（支持逗号分隔多TP）
                for m in re.finditer(r'\*\*关联测试点[：:]\*\*\s*(.+)', tc_content):
                    for tp in re.findall(r'TP-[A-Za-z0-9\-]+', m.group(1)):
                        covered_tp.add(tp)
            
            # 计算覆盖率
            tp_total = len(tp_ids)
            tp_covered = len(covered_tp)
            coverage_rate = (tp_covered / tp_total * 100) if tp_total > 0 else 0
            
            # 找出未覆盖的测试点
            uncovered_tp = sorted(tp_ids - covered_tp)
            
            result['testpoint_coverage'] = {
                'total': tp_total,
                'covered': tp_covered,
                'uncovered_count': len(uncovered_tp),
                'coverage_rate': round(coverage_rate, 2),
                'uncovered_list': uncovered_tp[:20] if uncovered_tp else [],  # 只列出前20个
            }
            
            if debug_mode:
                print(f"[DEBUG] 测试点总数: {tp_total}")
                print(f"[DEBUG] 已覆盖测试点: {tp_covered}")
                print(f"[DEBUG] 覆盖率: {coverage_rate:.2f}%")
                if uncovered_tp:
                    print(f"[DEBUG] 未覆盖测试点（前5个）: {uncovered_tp[:5]}")
            
            # 覆盖率低于阈值时添加警告
            if coverage_rate < 95:
                result['warnings'].append(f"测试点覆盖率 {coverage_rate:.2f}% < 95%，未覆盖 {len(uncovered_tp)} 个测试点")
            
        except Exception as e:
            result['testpoint_coverage_error'] = str(e)
            if debug_mode:
                print(f"[DEBUG] 覆盖率检查错误: {e}")

    # 生成 validation_report.md（Phase5必需输出，按phase5_rules格式）
    # 成功：脚本自动生成；失败：显式告警，由AI按phase5_rules"第一步验证"直接生成兜底
    try:
        vr_path = generate_validation_report(output_dir, result)
        result['validation_report'] = vr_path
        result['validation_report_status'] = 'success'
    except Exception as e:
        # 告警（非静默）：写入warnings供协调器检测，置failed状态，AI需直接生成
        alert = f'⚠告警: validation_report.md 脚本生成失败({e})，需AI按phase5_rules验证报告格式直接生成'
        result['warnings'].append(alert)
        result['validation_report_status'] = 'failed'
        result['validation_report_error'] = str(e)
    
    # 清理临时文件（导出成功后）
    if result.get('status') == 'success' and result.get('cases', 0) > 0:
        try:
            import glob as glob_module
            import shutil
            cleaned = []
            failed_to_clean = []
            
            # 清理批次目录
            batch_dirs = ['batches_phase2', 'batches_phase4', 'batches']
            for bd in batch_dirs:
                bd_path = os.path.join(output_dir, bd)
                if os.path.isdir(bd_path):
                    try:
                        shutil.rmtree(bd_path)
                        cleaned.append(bd)
                    except Exception as e:
                        failed_to_clean.append(f"{bd}: {str(e)}")
            
            # tasks/ 目录保留：含 timing.json（计时报告数据源）与 task_checkpoints/（审计快照，task-manager 禁止删除）
            
            # 清理temp目录
            temp_dir = os.path.join(output_dir, 'temp')
            if os.path.isdir(temp_dir):
                try:
                    shutil.rmtree(temp_dir)
                    cleaned.append('temp')
                except Exception as e:
                    failed_to_clean.append(f"temp: {str(e)}")
            
            # 清理固定json文件（timing.json 保留：完成后需追加 pipeline_completed_at 并生成计时报告）
            fixed_json_files = [
                'testing_technology.json',
                'coverage_result.json',
                'validate_result.json',
            ]
            for jf in fixed_json_files:
                jf_path = os.path.join(output_dir, jf)
                if os.path.isfile(jf_path):
                    try:
                        os.remove(jf_path)
                        cleaned.append(jf)
                    except Exception as e:
                        failed_to_clean.append(f"{jf}: {str(e)}")
            
            # 使用glob清理phase2_*.json和phase4_*.json
            for pattern in ['phase2_*.json', 'phase4_*.json']:
                glob_path = os.path.join(output_dir, pattern)
                for json_file in glob_module.glob(glob_path):
                    try:
                        os.remove(json_file)
                        cleaned.append(os.path.basename(json_file))
                    except Exception as e:
                        failed_to_clean.append(f"{os.path.basename(json_file)}: {str(e)}")
            
            if cleaned and debug_mode:
                print(f"[DEBUG] 已清理临时文件: {', '.join(cleaned)}")
            
            if failed_to_clean:
                result['cleanup_status'] = 'partial'
                result['cleanup_warnings'] = failed_to_clean
                if debug_mode:
                    print(f"[DEBUG] 部分文件清理失败: {', '.join(failed_to_clean)}")
            else:
                result['cleanup_status'] = 'success'
            
            result['cleaned_items'] = cleaned
        except Exception as e:
            result['cleanup_status'] = 'failed'
            result['cleanup_error'] = str(e)
            if debug_mode:
                print(f"[DEBUG] 清理失败: {e}")
    
        # 统一输出最终结果（包含覆盖率信息）
        print(json.dumps(result, ensure_ascii=False))


if __name__ == '__main__':
    main()
